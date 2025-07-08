from .helpers.Response import Response
from flask import ( Blueprint, request )
import openpyxl
from datetime import datetime
import os, time, pathlib
from flaskr.db import get_db
import string

bp = Blueprint('sheet', __name__, url_prefix='/sheets')

upload_sheets_dir = './uploads/sheets'
head_rows = 10


#
@bp.route('/upload', methods=['POST'])
def upload():
    uploaded_file = request.files['sheet']
    path = f"{upload_sheets_dir}/{int(time.time())}.xlsx";
    dir_exists = os.path.isdir(upload_sheets_dir)
    if  dir_exists == False:
        pathlib.Path(upload_sheets_dir).mkdir(parents=True)
    uploaded_file.save(path)

    sheet = read_sheet(path)

    didSucceed = True
    errors = []
    try:
        db = get_db()
        db.execute(
            'INSERT INTO sheet (name, path, rows, cols, created) VALUES (?, ?, ?, ?, ?)',
            (uploaded_file.filename, path, sheet.max_row, sheet.max_column, int(time.time()))
        )
        db.commit()
    except Exception as err:
        didSucceed = False;
        print(err);
        errors.append({ 'message': 'Unable to store file. Check logs for further info' });

    if didSucceed == False:
        jsonData = { 'success': didSucceed, 'message': 'Sheet upload failed', 'data': [], 'errors': errors }
        response = Response(jsonData, status = 422)
        return response.make_json_response()

    data = {
        'id': 2
    }

    jsonData = { 'success': didSucceed, 'message': 'Sheet uploaded', 'data': data, 'errors': [] }
    response = Response(jsonData, status = 200)
    return response.make_json_response()

@bp.route('/list', methods=['GET'])
def list():
    db = get_db()
    response = db.execute('SELECT * FROM sheet').fetchall()
    sheets = [];
    for x in response:
        sheets.append({
            'id': x['id'],
            'path': x['path'],
            'name': x['name'],
            'rows': x['rows'],
            'cols': x['cols'],
            'created': datetime.fromtimestamp(x['created']).strftime('%d-%b-%Y %H:%M'),
        })

    data = {
        'sheets': sheets
    }
    jsonData = { 'success': True, 'message': 'Sheets retrieved', 'data': data, 'errors': [] }
    response = Response(jsonData, status = 200)
    return response.make_json_response()

@bp.route('/list/<id>', methods=['GET'])
def show(id):
    db = get_db()
    response = db.execute( 'SELECT * FROM sheet WHERE id = ?', [id] ).fetchone()
    sheet = read_sheet(response['path'])
    head = read_sheet_head(sheet);

    columns = []
    for x in range(0, sheet.max_column):
        columns.append(string.ascii_uppercase[x])

    data = {
        'sheet': {
            'meta': {
                'id': response['id'],
                'name': response['name'],
                'rows': response['rows'],
                'cols': response['cols'],
            },
            'head': head,
            'columns': columns
        }
    }
    jsonData = { 'success': True, 'message': 'Sheets retrieved', 'data': data, 'errors': [] }
    response = Response(jsonData, status = 200)
    return response.make_json_response()

@bp.route('/<id>', methods=['DELETE'])
def trash(id):
    db = get_db()
    db.execute( 'DELETE FROM sheet WHERE id = ?', [id] );
    db.commit()
    jsonData = { 'success': True, 'message': 'Sheet upload failed', 'data': [], 'errors': [] }
    response = Response(jsonData, status = 200)
    return response.make_json_response()

def read_sheet(path):
    df = openpyxl.load_workbook(path)
    sheet = df.active
    return sheet

def read_sheet_head(sheet):
    rows = [];
    for row in range(0, head_rows):
        rowData = [];
        for col in sheet.iter_cols(1, sheet.max_column):
            rowData.append(col[row].value)
        rows.append(rowData)
    return rows;
